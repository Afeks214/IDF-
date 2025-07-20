"""
Excel Export Service with Hebrew RTL Support

Advanced Excel export service with comprehensive Hebrew support, RTL formatting,
and military-grade security features.
"""

import io
import os
from datetime import datetime
from typing import Dict, List, Optional, Any, Union
from pathlib import Path
import logging

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_GENERAL
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.protection import SheetProtection

from app.core.config import get_settings
from app.utils.hebrew import HebrewTextProcessor
from app.models.core import TestingData

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Advanced Excel Export Service with Hebrew RTL Support
    
    Features:
    - Hebrew text export with proper RTL formatting
    - Multi-sheet workbooks with complex layouts
    - Charts and graphs with Hebrew labels
    - Data validation and protection
    - Conditional formatting
    - Template-based exports
    - Military-grade security features
    """
    
    def __init__(self):
        self.settings = get_settings()
        self.hebrew_processor = HebrewTextProcessor()
        self._setup_default_styles()
        
    def _setup_default_styles(self):
        """Setup default Excel styles for Hebrew content"""
        # Hebrew font settings
        self.hebrew_font = Font(
            name='Arial',
            size=11,
            bold=False
        )
        
        self.hebrew_bold_font = Font(
            name='Arial',
            size=11,
            bold=True
        )
        
        self.hebrew_title_font = Font(
            name='Arial',
            size=14,
            bold=True
        )
        
        # RTL alignment
        self.rtl_alignment = Alignment(
            horizontal='right',
            vertical='center',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0,
            reading_order=2  # Right-to-left
        )
        
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0
        )
        
        # Color schemes
        self.header_fill = PatternFill(
            start_color='4472C4',
            end_color='4472C4',
            fill_type='solid'
        )
        
        self.alternate_fill = PatternFill(
            start_color='F2F2F2',
            end_color='F2F2F2',
            fill_type='solid'
        )
        
        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def export_inspection_data(
        self,
        inspection_data: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        template_name: Optional[str] = None
    ) -> bytes:
        """
        Export inspection data to Excel with Hebrew formatting
        
        Args:
            inspection_data: List of inspection records
            output_path: Optional file path to save Excel file
            template_name: Optional template name for custom formatting
            
        Returns:
            Excel file content as bytes
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create main data sheet
            self._create_main_data_sheet(wb, inspection_data)
            
            # Create summary sheet
            self._create_summary_sheet(wb, inspection_data)
            
            # Create charts sheet
            self._create_charts_sheet(wb, inspection_data)
            
            # Create pivot tables sheet
            self._create_pivot_sheet(wb, inspection_data)
            
            # Apply template if specified
            if template_name:
                self._apply_template(wb, template_name)
                
            # Set workbook properties
            self._set_workbook_properties(wb)
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            # Save to file if path provided
            if output_path:
                with open(output_path, 'wb') as f:
                    f.write(excel_content)
                    
            logger.info(f"Exported {len(inspection_data)} inspection records to Excel")
            return excel_content
            
        except Exception as e:
            logger.error(f"Error exporting inspection data: {e}")
            raise
            
    def _create_main_data_sheet(self, wb: Workbook, data: List[Dict[str, Any]]):
        """Create main data sheet with inspection records"""
        ws = wb.create_sheet("נתוני בדיקות", 0)
        
        # Hebrew column headers
        headers = [
            'מספר מערכת',
            'תאריך בדיקה',
            'שם בודק',
            'סטטוס',
            'ציון',
            'הערות',
            'מיקום',
            'קטגוריה',
            'עדיפות',
            'תאריך עדכון'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.hebrew_bold_font
            cell.alignment = self.rtl_alignment
            cell.fill = self.header_fill
            cell.border = self.thin_border
            
        # Write data rows
        for row_idx, record in enumerate(data, 2):
            row_data = [
                record.get('system_id', ''),
                record.get('inspection_date', ''),
                record.get('inspector_name', ''),
                record.get('status', ''),
                record.get('score', ''),
                record.get('notes', ''),
                record.get('location', ''),
                record.get('category', ''),
                record.get('priority', ''),
                record.get('updated_at', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = self.hebrew_font
                cell.alignment = self.rtl_alignment
                cell.border = self.thin_border
                
                # Alternate row coloring
                if row_idx % 2 == 0:
                    cell.fill = self.alternate_fill
                    
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Add data validation
        self._add_data_validation(ws, len(data))
        
        # Add conditional formatting
        self._add_conditional_formatting(ws, len(data))
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict[str, Any]]):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("סיכום", 1)
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="סיכום בדיקות")
        title_cell.font = self.hebrew_title_font
        title_cell.alignment = self.center_alignment
        ws.merge_cells('A1:D1')
        
        # Calculate statistics
        total_inspections = len(data)
        passed_inspections = sum(1 for item in data if item.get('status') == 'עבר')
        failed_inspections = total_inspections - passed_inspections
        success_rate = (passed_inspections / total_inspections) * 100 if total_inspections > 0 else 0
        
        # Summary statistics
        summary_data = [
            ['סך הכל בדיקות', total_inspections],
            ['בדיקות שעברו', passed_inspections],
            ['בדיקות שנכשלו', failed_inspections],
            ['אחוז הצלחה', f"{success_rate:.1f}%"]
        ]
        
        # Write summary data
        for row_idx, (label, value) in enumerate(summary_data, 3):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            label_cell.font = self.hebrew_bold_font
            label_cell.alignment = self.rtl_alignment
            label_cell.border = self.thin_border
            
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.font = self.hebrew_font
            value_cell.alignment = self.center_alignment
            value_cell.border = self.thin_border
            
        # Category breakdown
        category_stats = self._calculate_category_stats(data)
        
        # Write category breakdown
        ws.cell(row=8, column=1, value="פילוח לפי קטגוריה").font = self.hebrew_bold_font
        
        for row_idx, (category, count) in enumerate(category_stats.items(), 9):
            cat_cell = ws.cell(row=row_idx, column=1, value=category)
            cat_cell.font = self.hebrew_font
            cat_cell.alignment = self.rtl_alignment
            cat_cell.border = self.thin_border
            
            count_cell = ws.cell(row=row_idx, column=2, value=count)
            count_cell.font = self.hebrew_font
            count_cell.alignment = self.center_alignment
            count_cell.border = self.thin_border
            
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
    def _create_charts_sheet(self, wb: Workbook, data: List[Dict[str, Any]]):
        """Create charts and visualizations sheet"""
        ws = wb.create_sheet("גרפים", 2)
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="גרפים ותרשימים")
        title_cell.font = self.hebrew_title_font
        title_cell.alignment = self.center_alignment
        
        # Create pie chart for status distribution
        self._create_status_pie_chart(ws, data)
        
        # Create bar chart for category breakdown
        self._create_category_bar_chart(ws, data)
        
        # Create line chart for trends
        self._create_trend_line_chart(ws, data)
        
    def _create_pivot_sheet(self, wb: Workbook, data: List[Dict[str, Any]]):
        """Create pivot tables sheet"""
        ws = wb.create_sheet("טבלאות מרכזות", 3)
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="טבלאות מרכזות")
        title_cell.font = self.hebrew_title_font
        title_cell.alignment = self.center_alignment
        
        # Create pivot table data
        pivot_data = self._create_pivot_table_data(data)
        
        # Write pivot table
        for row_idx, row_data in enumerate(pivot_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.hebrew_font
                cell.alignment = self.rtl_alignment
                cell.border = self.thin_border
                
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
    def _calculate_category_stats(self, data: List[Dict[str, Any]]) -> Dict[str, int]:
        """Calculate statistics by category"""
        category_stats = {}
        
        for record in data:
            category = record.get('category', 'לא מוגדר')
            category_stats[category] = category_stats.get(category, 0) + 1
            
        return category_stats
        
    def _create_status_pie_chart(self, ws, data: List[Dict[str, Any]]):
        """Create pie chart for status distribution"""
        try:
            # Status distribution data
            status_counts = {}
            for record in data:
                status = record.get('status', 'לא מוגדר')
                status_counts[status] = status_counts.get(status, 0) + 1
                
            # Write chart data
            chart_data_start_row = 5
            ws.cell(row=chart_data_start_row, column=5, value="סטטוס")
            ws.cell(row=chart_data_start_row, column=6, value="כמות")
            
            for idx, (status, count) in enumerate(status_counts.items(), 1):
                ws.cell(row=chart_data_start_row + idx, column=5, value=status)
                ws.cell(row=chart_data_start_row + idx, column=6, value=count)
                
            # Create pie chart
            pie_chart = PieChart()
            pie_chart.title = "התפלגות סטטוס"
            
            # Set data range
            labels = Reference(ws, min_col=5, min_row=chart_data_start_row + 1, 
                             max_row=chart_data_start_row + len(status_counts))
            data_ref = Reference(ws, min_col=6, min_row=chart_data_start_row + 1,
                               max_row=chart_data_start_row + len(status_counts))
            
            pie_chart.add_data(data_ref, titles_from_data=False)
            pie_chart.set_categories(labels)
            
            # Add chart to worksheet
            ws.add_chart(pie_chart, "H5")
            
        except Exception as e:
            logger.error(f"Error creating pie chart: {e}")
            
    def _create_category_bar_chart(self, ws, data: List[Dict[str, Any]]):
        """Create bar chart for category breakdown"""
        try:
            category_stats = self._calculate_category_stats(data)
            
            # Write chart data
            chart_data_start_row = 15
            ws.cell(row=chart_data_start_row, column=5, value="קטגוריה")
            ws.cell(row=chart_data_start_row, column=6, value="כמות")
            
            for idx, (category, count) in enumerate(category_stats.items(), 1):
                ws.cell(row=chart_data_start_row + idx, column=5, value=category)
                ws.cell(row=chart_data_start_row + idx, column=6, value=count)
                
            # Create bar chart
            bar_chart = BarChart()
            bar_chart.title = "פילוח לפי קטגוריה"
            bar_chart.x_axis.title = "קטגוריה"
            bar_chart.y_axis.title = "כמות"
            
            # Set data range
            labels = Reference(ws, min_col=5, min_row=chart_data_start_row + 1,
                             max_row=chart_data_start_row + len(category_stats))
            data_ref = Reference(ws, min_col=6, min_row=chart_data_start_row + 1,
                               max_row=chart_data_start_row + len(category_stats))
            
            bar_chart.add_data(data_ref, titles_from_data=False)
            bar_chart.set_categories(labels)
            
            # Add chart to worksheet
            ws.add_chart(bar_chart, "H15")
            
        except Exception as e:
            logger.error(f"Error creating bar chart: {e}")
            
    def _create_trend_line_chart(self, ws, data: List[Dict[str, Any]]):
        """Create line chart for trends"""
        try:
            # Group data by date
            date_stats = {}
            for record in data:
                date = record.get('inspection_date', '')
                if date:
                    date_stats[date] = date_stats.get(date, 0) + 1
                    
            # Sort by date
            sorted_dates = sorted(date_stats.items())
            
            # Write chart data
            chart_data_start_row = 25
            ws.cell(row=chart_data_start_row, column=5, value="תאריך")
            ws.cell(row=chart_data_start_row, column=6, value="מספר בדיקות")
            
            for idx, (date, count) in enumerate(sorted_dates, 1):
                ws.cell(row=chart_data_start_row + idx, column=5, value=date)
                ws.cell(row=chart_data_start_row + idx, column=6, value=count)
                
            # Create line chart
            line_chart = LineChart()
            line_chart.title = "מגמת בדיקות לאורך זמן"
            line_chart.x_axis.title = "תאריך"
            line_chart.y_axis.title = "מספר בדיקות"
            
            # Set data range
            labels = Reference(ws, min_col=5, min_row=chart_data_start_row + 1,
                             max_row=chart_data_start_row + len(sorted_dates))
            data_ref = Reference(ws, min_col=6, min_row=chart_data_start_row + 1,
                               max_row=chart_data_start_row + len(sorted_dates))
            
            line_chart.add_data(data_ref, titles_from_data=False)
            line_chart.set_categories(labels)
            
            # Add chart to worksheet
            ws.add_chart(line_chart, "H25")
            
        except Exception as e:
            logger.error(f"Error creating line chart: {e}")
            
    def _create_pivot_table_data(self, data: List[Dict[str, Any]]) -> List[List[str]]:
        """Create pivot table data"""
        pivot_data = []
        
        # Headers
        pivot_data.append(['קטגוריה', 'סטטוס', 'כמות'])
        
        # Group by category and status
        category_status_counts = {}
        for record in data:
            category = record.get('category', 'לא מוגדר')
            status = record.get('status', 'לא מוגדר')
            key = (category, status)
            category_status_counts[key] = category_status_counts.get(key, 0) + 1
            
        # Add data rows
        for (category, status), count in category_status_counts.items():
            pivot_data.append([category, status, str(count)])
            
        return pivot_data
        
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _add_data_validation(self, ws, row_count: int):
        """Add data validation to specific columns"""
        try:
            # Status column validation
            status_validation = DataValidation(
                type="list",
                formula1='"עבר,נכשל,בהמתנה"',
                allow_blank=True
            )
            ws.add_data_validation(status_validation)
            status_validation.add(f'D2:D{row_count + 1}')
            
            # Priority column validation
            priority_validation = DataValidation(
                type="list",
                formula1='"גבוה,בינוני,נמוך"',
                allow_blank=True
            )
            ws.add_data_validation(priority_validation)
            priority_validation.add(f'I2:I{row_count + 1}')
            
        except Exception as e:
            logger.error(f"Error adding data validation: {e}")
            
    def _add_conditional_formatting(self, ws, row_count: int):
        """Add conditional formatting for better visualization"""
        try:
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
            from openpyxl.styles import PatternFill
            
            # Status column conditional formatting
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            
            # Failed status - red
            failed_rule = CellIsRule(operator='equal', formula=['"נכשל"'], fill=red_fill)
            ws.conditional_formatting.add(f'D2:D{row_count + 1}', failed_rule)
            
            # Passed status - green
            passed_rule = CellIsRule(operator='equal', formula=['"עבר"'], fill=green_fill)
            ws.conditional_formatting.add(f'D2:D{row_count + 1}', passed_rule)
            
            # Pending status - yellow
            pending_rule = CellIsRule(operator='equal', formula=['"בהמתנה"'], fill=yellow_fill)
            ws.conditional_formatting.add(f'D2:D{row_count + 1}', pending_rule)
            
        except Exception as e:
            logger.error(f"Error adding conditional formatting: {e}")
            
    def _apply_template(self, wb: Workbook, template_name: str):
        """Apply predefined template formatting"""
        try:
            # This would load template configurations
            # For now, apply default military template
            
            # Add security notice
            security_sheet = wb.create_sheet("הודעת אבטחה", 0)
            security_notice = """
            מסמך זה מכיל מידע רגיש וסודי.
            אין להעביר או לשתף מידע זה ללא אישור מתאים.
            כל השימושים במסמך זה מתועדים ומבוקרים.
            """
            
            security_cell = security_sheet.cell(row=2, column=1, value=security_notice)
            security_cell.font = self.hebrew_font
            security_cell.alignment = self.rtl_alignment
            security_sheet.merge_cells('A2:D10')
            
            # Set sheet protection
            for sheet in wb.worksheets:
                if sheet.title != "הודעת אבטחה":
                    sheet.protection = SheetProtection(
                        sheet=True,
                        objects=True,
                        scenarios=True,
                        formatCells=False,
                        formatColumns=False,
                        formatRows=False,
                        insertColumns=False,
                        insertRows=False,
                        deleteColumns=False,
                        deleteRows=False
                    )
                    
        except Exception as e:
            logger.error(f"Error applying template: {e}")
            
    def _set_workbook_properties(self, wb: Workbook):
        """Set workbook metadata and properties"""
        try:
            wb.properties.title = "דוח בדיקות IDF"
            wb.properties.subject = "דוח בדיקות מערכות"
            wb.properties.creator = "מערכת IDF"
            wb.properties.description = "דוח מפורט של בדיקות מערכות עם תמיכה בעברית"
            wb.properties.language = "he-IL"
            wb.properties.created = datetime.now()
            wb.properties.modified = datetime.now()
            
        except Exception as e:
            logger.error(f"Error setting workbook properties: {e}")
            
    def export_summary_report(
        self,
        summary_data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> bytes:
        """Export summary report to Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "דוח סיכום"
            
            # Title
            title_cell = ws.cell(row=1, column=1, value="דוח סיכום מנהלים")
            title_cell.font = self.hebrew_title_font
            title_cell.alignment = self.center_alignment
            ws.merge_cells('A1:D1')
            
            # Add summary data
            for row_idx, (key, value) in enumerate(summary_data.items(), 3):
                key_cell = ws.cell(row=row_idx, column=1, value=key)
                key_cell.font = self.hebrew_bold_font
                key_cell.alignment = self.rtl_alignment
                
                value_cell = ws.cell(row=row_idx, column=2, value=value)
                value_cell.font = self.hebrew_font
                value_cell.alignment = self.center_alignment
                
            # Auto-adjust columns
            self._auto_adjust_columns(ws)
            
            # Set workbook properties
            self._set_workbook_properties(wb)
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            if output_path:
                with open(output_path, 'wb') as f:
                    f.write(excel_content)
                    
            return excel_content
            
        except Exception as e:
            logger.error(f"Error exporting summary report: {e}")
            raise
            
    def export_custom_report(
        self,
        data: Dict[str, Any],
        template_config: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> bytes:
        """Export custom report based on template configuration"""
        try:
            wb = Workbook()
            
            # Process template sections
            for section_name, section_config in template_config.get('sections', {}).items():
                ws = wb.create_sheet(section_name)
                self._build_custom_sheet(ws, section_config, data)
                
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                
            # Set workbook properties
            self._set_workbook_properties(wb)
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            if output_path:
                with open(output_path, 'wb') as f:
                    f.write(excel_content)
                    
            return excel_content
            
        except Exception as e:
            logger.error(f"Error exporting custom report: {e}")
            raise
            
    def _build_custom_sheet(self, ws, section_config: Dict[str, Any], data: Dict[str, Any]):
        """Build custom sheet based on section configuration"""
        try:
            # Set sheet title
            ws.title = section_config.get('title', 'דף מותאם')
            
            # Add headers if specified
            if 'headers' in section_config:
                for col, header in enumerate(section_config['headers'], 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = self.hebrew_bold_font
                    cell.alignment = self.rtl_alignment
                    cell.fill = self.header_fill
                    cell.border = self.thin_border
                    
            # Add data based on configuration
            data_source = section_config.get('data_source', '')
            if data_source in data:
                section_data = data[data_source]
                
                if isinstance(section_data, list):
                    for row_idx, record in enumerate(section_data, 2):
                        if isinstance(record, dict):
                            for col, field in enumerate(section_config.get('fields', []), 1):
                                cell = ws.cell(row=row_idx, column=col, value=record.get(field, ''))
                                cell.font = self.hebrew_font
                                cell.alignment = self.rtl_alignment
                                cell.border = self.thin_border
                                
            # Auto-adjust columns
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error building custom sheet: {e}")
            
    def create_template_file(self, template_name: str, template_config: Dict[str, Any]) -> str:
        """Create Excel template file for reuse"""
        try:
            template_path = Path(__file__).parent / "templates" / f"{template_name}.xlsx"
            template_path.parent.mkdir(exist_ok=True)
            
            wb = Workbook()
            
            # Build template based on configuration
            for sheet_name, sheet_config in template_config.get('sheets', {}).items():
                ws = wb.create_sheet(sheet_name)
                self._build_template_sheet(ws, sheet_config)
                
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                
            # Save template
            wb.save(template_path)
            
            logger.info(f"Created template file: {template_path}")
            return str(template_path)
            
        except Exception as e:
            logger.error(f"Error creating template file: {e}")
            raise
            
    def _build_template_sheet(self, ws, sheet_config: Dict[str, Any]):
        """Build template sheet structure"""
        try:
            # Add headers
            if 'headers' in sheet_config:
                for col, header in enumerate(sheet_config['headers'], 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = self.hebrew_bold_font
                    cell.alignment = self.rtl_alignment
                    cell.fill = self.header_fill
                    cell.border = self.thin_border
                    
            # Add sample data if specified
            if 'sample_data' in sheet_config:
                for row_idx, sample_row in enumerate(sheet_config['sample_data'], 2):
                    for col, value in enumerate(sample_row, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.font = self.hebrew_font
                        cell.alignment = self.rtl_alignment
                        cell.border = self.thin_border
                        
            # Auto-adjust columns
            self._auto_adjust_columns(ws)
            
        except Exception as e:
            logger.error(f"Error building template sheet: {e}")