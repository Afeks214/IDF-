"""
Excel import/export service for IDF Testing Infrastructure.
Handles Hebrew text preservation and data validation.
"""

import pandas as pd
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import logging
from datetime import datetime, date
import hashlib
import json

from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..models.core import Inspection, Building, InspectionType, Regulator
from ..models.base import BaseModel

logger = logging.getLogger(__name__)


class ExcelImportService:
    """Service for importing Excel data with Hebrew text support."""
    
    def __init__(self, session: AsyncSession):
        self.session = session
        self.stats = {
            "processed": 0,
            "imported": 0,
            "skipped": 0,
            "errors": 0
        }
    
    async def import_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Import Excel file with Hebrew data.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Import statistics and results
        """
        try:
            # Load Excel file with proper encoding
            workbook = load_workbook(file_path, data_only=True)
            
            results = {
                "worksheets": {},
                "stats": self.stats,
                "errors": []
            }
            
            # Process main data worksheet (טבלה מרכזת)
            if "טבלה מרכזת" in workbook.sheetnames:
                main_results = await self._import_main_data(workbook["טבלה מרכזת"])
                results["worksheets"]["main_data"] = main_results
            
            # Process lookup data (ערכים)
            if "ערכים" in workbook.sheetnames:
                lookup_results = await self._import_lookup_data(workbook["ערכים"])
                results["worksheets"]["lookup_data"] = lookup_results
            
            await self.session.commit()
            return results
            
        except Exception as e:
            await self.session.rollback()
            logger.error(f"Excel import failed: {str(e)}")
            raise
    
    async def _import_main_data(self, worksheet) -> Dict[str, Any]:
        """Import main inspection data from worksheet."""
        
        results = {"imported": 0, "errors": []}
        
        # Get all rows with data
        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        
        for row_idx, row in enumerate(rows, start=2):
            try:
                if not any(row):  # Skip empty rows
                    continue
                
                # Map Excel columns to database fields
                inspection_data = self._map_excel_row_to_inspection(row)
                
                if inspection_data:
                    # Create or update inspection record
                    inspection = Inspection(**inspection_data)
                    
                    # Sanitize Hebrew text
                    for field in ["building_manager", "red_team", "inspection_type", 
                                "inspection_leader", "inspection_notes"]:
                        if hasattr(inspection, field):
                            value = getattr(inspection, field)
                            if isinstance(value, str):
                                sanitized = BaseModel.sanitize_hebrew_text(value)
                                normalized = BaseModel.normalize_hebrew_text(sanitized)
                                setattr(inspection, field, normalized)
                    
                    self.session.add(inspection)
                    results["imported"] += 1
                    self.stats["imported"] += 1
                
                self.stats["processed"] += 1
                
            except Exception as e:
                error_msg = f"Row {row_idx}: {str(e)}"
                results["errors"].append(error_msg)
                self.stats["errors"] += 1
                logger.error(error_msg)
        
        return results
    
    def _map_excel_row_to_inspection(self, row: Tuple) -> Optional[Dict[str, Any]]:
        """Map Excel row to inspection data structure."""
        
        if len(row) < 18:  # Minimum required columns
            return None
        
        try:
            # Column mapping based on Excel structure
            data = {
                "building_id": self._clean_text(row[0]),
                "building_manager": self._clean_text(row[1]),
                "red_team": self._clean_text(row[2]),
                "inspection_type": self._clean_text(row[3]),
                "inspection_leader": self._clean_text(row[4]),
                "inspection_round": self._parse_int(row[5]),
                "regulator_1": self._clean_text(row[6]),
                "regulator_2": self._clean_text(row[7]),
                "regulator_3": self._clean_text(row[8]),
                "regulator_4": self._clean_text(row[9]),
                "execution_schedule": self._parse_date(row[10]),
                "target_completion": self._parse_date(row[11]),
                "coordinated_with_contractor": self._parse_boolean(row[12]),
                "defect_report_path": self._clean_text(row[13]),
                "report_distributed": self._parse_boolean(row[14]),
                "distribution_date": self._parse_date(row[15]),
                "repeat_inspection": self._parse_boolean(row[16]),
                "inspection_notes": self._clean_text(row[17]) if len(row) > 17 else None,
            }
            
            # Remove None values and empty strings
            return {k: v for k, v in data.items() if v is not None and v != ""}
            
        except Exception as e:
            logger.error(f"Error mapping row: {str(e)}")
            return None
    
    async def _import_lookup_data(self, worksheet) -> Dict[str, Any]:
        """Import lookup data from values worksheet."""
        
        results = {"buildings": 0, "types": 0, "regulators": 0, "errors": []}
        
        try:
            # Process different sections of lookup data
            # This would need to be customized based on actual worksheet structure
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            
            for row in rows:
                if not any(row):
                    continue
                
                # Example: Import building data if found in specific columns
                if row[0] and row[1]:  # Building code and name
                    building_data = {
                        "building_code": self._clean_text(row[0]),
                        "building_name": self._clean_text(row[1]),
                        "manager_name": self._clean_text(row[2]) if len(row) > 2 else None,
                    }
                    
                    # Check if building already exists
                    existing = await self.session.get(Building, building_data["building_code"])
                    if not existing:
                        building = Building(**building_data)
                        self.session.add(building)
                        results["buildings"] += 1
            
            return results
            
        except Exception as e:
            logger.error(f"Lookup data import error: {str(e)}")
            results["errors"].append(str(e))
            return results
    
    def _clean_text(self, value: Any) -> Optional[str]:
        """Clean and normalize text values."""
        if value is None:
            return None
        
        text = str(value).strip()
        if not text or text.lower() in ["nan", "none", "null", ""]:
            return None
        
        # Apply Hebrew text normalization
        return BaseModel.normalize_hebrew_text(text)
    
    def _parse_int(self, value: Any) -> Optional[int]:
        """Parse integer values safely."""
        if value is None:
            return None
        
        try:
            if isinstance(value, (int, float)):
                return int(value)
            
            text = str(value).strip()
            if not text or text.lower() in ["nan", "none", "null", ""]:
                return None
            
            return int(float(text))
        except:
            return None
    
    def _parse_date(self, value: Any) -> Optional[date]:
        """Parse date values safely."""
        if value is None:
            return None
        
        try:
            if isinstance(value, datetime):
                return value.date()
            elif isinstance(value, date):
                return value
            
            text = str(value).strip()
            if not text or text.lower() in ["nan", "none", "null", ""]:
                return None
            
            # Try parsing common date formats
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]:
                try:
                    return datetime.strptime(text, fmt).date()
                except:
                    continue
            
            return None
        except:
            return None
    
    def _parse_boolean(self, value: Any) -> bool:
        """Parse boolean values safely."""
        if value is None:
            return False
        
        if isinstance(value, bool):
            return value
        
        text = str(value).strip().lower()
        
        # Hebrew yes/no values
        if text in ["כן", "yes", "true", "1", "v", "✓"]:
            return True
        elif text in ["לא", "no", "false", "0", "x", "✗"]:
            return False
        
        return False


class ExcelExportService:
    """Service for exporting data to Excel with Hebrew support."""
    
    def __init__(self, session: AsyncSession):
        self.session = session
    
    async def export_inspections_to_excel(self, file_path: str, filters: Optional[Dict] = None) -> str:
        """
        Export inspections to Excel file with Hebrew formatting.
        
        Args:
            file_path: Output file path
            filters: Optional filters for data selection
            
        Returns:
            Path to created file
        """
        try:
            # Query inspections with filters
            query = self.session.query(Inspection)
            
            if filters:
                if "building_id" in filters:
                    query = query.filter(Inspection.building_id == filters["building_id"])
                if "status" in filters:
                    query = query.filter(Inspection.status == filters["status"])
                if "date_from" in filters:
                    query = query.filter(Inspection.execution_schedule >= filters["date_from"])
                if "date_to" in filters:
                    query = query.filter(Inspection.execution_schedule <= filters["date_to"])
            
            inspections = await query.all()
            
            # Convert to DataFrame
            data = []
            for inspection in inspections:
                row = {
                    "מבנה": inspection.building_id,
                    "מנהל מבנה": inspection.building_manager,
                    "צוות אדום": inspection.red_team,
                    "סוג הבדיקה": inspection.inspection_type,
                    "מוביל הבדיקה": inspection.inspection_leader,
                    "סבב בדיקות": inspection.inspection_round,
                    "רגולטור 1": inspection.regulator_1,
                    "רגולטור 2": inspection.regulator_2,
                    "רגולטור 3": inspection.regulator_3,
                    "רגולטור 4": inspection.regulator_4,
                    "לוז ביצוע": inspection.execution_schedule,
                    "יעד לסיום": inspection.target_completion,
                    "מתואם מול זכיין": "כן" if inspection.coordinated_with_contractor else "לא",
                    "נתיב דוח ליקויים": inspection.defect_report_path,
                    "דוח הופץ": "כן" if inspection.report_distributed else "לא",
                    "תאריך הפצה": inspection.distribution_date,
                    "בדיקה חוזרת": "כן" if inspection.repeat_inspection else "לא",
                    "התרשמות": inspection.inspection_notes,
                    "סטטוס": inspection.get_hebrew_status(),
                }
                data.append(row)
            
            df = pd.DataFrame(data)
            
            # Create Excel file with Hebrew formatting
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='בדיקות', index=False, encoding='utf-8')
                
                # Apply Hebrew formatting
                workbook = writer.book
                worksheet = writer.sheets['בדיקות']
                
                # Set RTL direction
                worksheet.sheet_view.rightToLeft = True
                
                # Format headers
                header_font = Font(bold=True, name='Arial', size=12)
                header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Excel export completed: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise