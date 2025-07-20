"""
Enhanced Hebrew Data Processor Service for IDF Testing Infrastructure.
Supports multi-format data ingestion, streaming processing, and quality validation.
"""

import asyncio
import json
import logging
import pandas as pd
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Any, AsyncIterator, Union, Tuple
from enum import Enum
import unicodedata
import re
from dataclasses import dataclass, field
from collections import defaultdict
import hashlib
import mimetypes

from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from openpyxl.cell import Cell
import chardet
import aiofiles
import aiocsv

from ..models.base import BaseModel
from ..models.core import Inspection, Building, InspectionType, Regulator
from ..core.validation import ValidationService
from ..utils.logger import get_logger

logger = get_logger(__name__)


class DataFormat(Enum):
    """Supported data formats for processing."""
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"
    TEXT = "text"
    UNKNOWN = "unknown"


class ProcessingStatus(Enum):
    """Processing status for data records."""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    SKIPPED = "skipped"


@dataclass
class DataQualityMetrics:
    """Data quality metrics for monitoring."""
    total_records: int = 0
    valid_records: int = 0
    invalid_records: int = 0
    hebrew_text_issues: int = 0
    encoding_issues: int = 0
    format_issues: int = 0
    duplicate_records: int = 0
    missing_required_fields: int = 0
    quality_score: float = 0.0
    issues: List[str] = field(default_factory=list)


@dataclass
class ProcessingResult:
    """Result of data processing operation."""
    status: ProcessingStatus
    records_processed: int
    records_imported: int
    records_skipped: int
    records_failed: int
    quality_metrics: DataQualityMetrics
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    processing_time: float = 0.0
    file_hash: Optional[str] = None


class HebrewDataProcessor:
    """Enhanced data processor with Hebrew text support and streaming capabilities."""
    
    # Hebrew text patterns
    HEBREW_LETTERS = re.compile(r'[\u05D0-\u05EA]')
    HEBREW_POINTS = re.compile(r'[\u05B0-\u05BD\u05BF-\u05C2\u05C4-\u05C5\u05C7]')
    HEBREW_BLOCK = re.compile(r'[\u0590-\u05FF]')
    
    # Data validation patterns
    PHONE_PATTERN = re.compile(r'^[\d\-\(\)\s\+]+$')
    EMAIL_PATTERN = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    def __init__(self, session: AsyncSession, chunk_size: int = 1000):
        """
        Initialize the Hebrew data processor.
        
        Args:
            session: Database session
            chunk_size: Number of records to process in each chunk
        """
        self.session = session
        self.chunk_size = chunk_size
        self.validation_service = ValidationService()
        self._processed_hashes = set()
        
    async def process_file(self, file_path: str, format_hint: Optional[DataFormat] = None) -> ProcessingResult:
        """
        Process a data file with automatic format detection.
        
        Args:
            file_path: Path to the file to process
            format_hint: Optional format hint to skip auto-detection
            
        Returns:
            Processing result with metrics and status
        """
        start_time = datetime.now()
        
        try:
            # Detect file format
            file_format = format_hint or await self._detect_format(file_path)
            
            # Calculate file hash for duplicate detection
            file_hash = await self._calculate_file_hash(file_path)
            
            logger.info(f"Processing file {file_path} as {file_format.value}, hash: {file_hash}")
            
            # Route to appropriate processor
            if file_format == DataFormat.EXCEL:
                result = await self._process_excel_file(file_path, file_hash)
            elif file_format == DataFormat.CSV:
                result = await self._process_csv_file(file_path, file_hash)
            elif file_format == DataFormat.JSON:
                result = await self._process_json_file(file_path, file_hash)
            else:
                raise ValueError(f"Unsupported format: {file_format}")
            
            # Calculate processing time
            processing_time = (datetime.now() - start_time).total_seconds()
            result.processing_time = processing_time
            
            logger.info(f"File processing completed in {processing_time:.2f}s: {result.records_processed} records")
            return result
            
        except Exception as e:
            logger.error(f"File processing failed: {str(e)}")
            return ProcessingResult(
                status=ProcessingStatus.FAILED,
                records_processed=0,
                records_imported=0,
                records_skipped=0,
                records_failed=0,
                quality_metrics=DataQualityMetrics(),
                errors=[str(e)],
                processing_time=(datetime.now() - start_time).total_seconds()
            )
    
    async def stream_process_data(self, data_stream: AsyncIterator[Dict[str, Any]]) -> AsyncIterator[ProcessingResult]:
        """
        Stream process data records with real-time quality validation.
        
        Args:
            data_stream: Async iterator of data records
            
        Yields:
            Processing results for each chunk
        """
        chunk = []
        chunk_number = 0
        
        async for record in data_stream:
            chunk.append(record)
            
            if len(chunk) >= self.chunk_size:
                chunk_number += 1
                logger.debug(f"Processing chunk {chunk_number} with {len(chunk)} records")
                
                result = await self._process_data_chunk(chunk, chunk_number)
                yield result
                
                chunk = []
        
        # Process remaining records
        if chunk:
            chunk_number += 1
            result = await self._process_data_chunk(chunk, chunk_number)
            yield result
    
    async def validate_data_quality(self, data: List[Dict[str, Any]]) -> DataQualityMetrics:
        """
        Validate data quality with Hebrew text analysis.
        
        Args:
            data: List of data records to validate
            
        Returns:
            Data quality metrics
        """
        metrics = DataQualityMetrics()
        metrics.total_records = len(data)
        
        seen_hashes = set()
        required_fields = ["building_id", "inspection_type"]
        
        for record in data:
            is_valid = True
            
            # Check for required fields
            missing_fields = [field for field in required_fields if not record.get(field)]
            if missing_fields:
                metrics.missing_required_fields += 1
                metrics.issues.append(f"Missing required fields: {', '.join(missing_fields)}")
                is_valid = False
            
            # Check for duplicates
            record_hash = self._calculate_record_hash(record)
            if record_hash in seen_hashes:
                metrics.duplicate_records += 1
                metrics.issues.append(f"Duplicate record found: {record_hash}")
                is_valid = False
            seen_hashes.add(record_hash)
            
            # Validate Hebrew text fields
            hebrew_issues = await self._validate_hebrew_fields(record)
            if hebrew_issues:
                metrics.hebrew_text_issues += 1
                metrics.issues.extend(hebrew_issues)
                is_valid = False
            
            # Validate data formats
            format_issues = await self._validate_data_formats(record)
            if format_issues:
                metrics.format_issues += 1
                metrics.issues.extend(format_issues)
                is_valid = False
            
            if is_valid:
                metrics.valid_records += 1
            else:
                metrics.invalid_records += 1
        
        # Calculate quality score
        if metrics.total_records > 0:
            metrics.quality_score = metrics.valid_records / metrics.total_records
        
        return metrics
    
    async def _detect_format(self, file_path: str) -> DataFormat:
        """Detect file format based on extension and content."""
        path = Path(file_path)
        extension = path.suffix.lower()
        
        # Check by extension first
        if extension in ['.xlsx', '.xls']:
            return DataFormat.EXCEL
        elif extension == '.csv':
            return DataFormat.CSV
        elif extension == '.json':
            return DataFormat.JSON
        
        # Check by MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type:
            if 'spreadsheet' in mime_type or 'excel' in mime_type:
                return DataFormat.EXCEL
            elif 'csv' in mime_type:
                return DataFormat.CSV
            elif 'json' in mime_type:
                return DataFormat.JSON
        
        # Check by content (first few bytes)
        try:
            async with aiofiles.open(file_path, 'rb') as f:
                header = await f.read(512)
                
                # Excel files start with PK (ZIP header)
                if header.startswith(b'PK'):
                    return DataFormat.EXCEL
                
                # Try to decode as text
                try:
                    decoded = header.decode('utf-8')
                    if decoded.strip().startswith('{') or decoded.strip().startswith('['):
                        return DataFormat.JSON
                    elif ',' in decoded and '\n' in decoded:
                        return DataFormat.CSV
                except UnicodeDecodeError:
                    pass
        except Exception:
            pass
        
        return DataFormat.UNKNOWN
    
    async def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of file content."""
        hash_sha256 = hashlib.sha256()
        
        async with aiofiles.open(file_path, 'rb') as f:
            while chunk := await f.read(8192):
                hash_sha256.update(chunk)
        
        return hash_sha256.hexdigest()
    
    def _calculate_record_hash(self, record: Dict[str, Any]) -> str:
        """Calculate hash of record for duplicate detection."""
        # Sort keys for consistent hashing
        sorted_record = {k: v for k, v in sorted(record.items()) if v is not None}
        record_str = json.dumps(sorted_record, sort_keys=True, ensure_ascii=False)
        return hashlib.md5(record_str.encode('utf-8')).hexdigest()
    
    async def _process_excel_file(self, file_path: str, file_hash: str) -> ProcessingResult:
        """Process Excel file with Hebrew text support."""
        result = ProcessingResult(
            status=ProcessingStatus.PROCESSING,
            records_processed=0,
            records_imported=0,
            records_skipped=0,
            records_failed=0,
            quality_metrics=DataQualityMetrics(),
            file_hash=file_hash
        )
        
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            # Process main worksheet
            main_sheet = None
            for sheet_name in ["טבלה מרכזת", "Main Data", "Sheet1"]:
                if sheet_name in workbook.sheetnames:
                    main_sheet = workbook[sheet_name]
                    break
            
            if not main_sheet:
                raise ValueError("No main data sheet found")
            
            # Convert to data records
            data_records = []
            headers = None
            
            for row_idx, row in enumerate(main_sheet.iter_rows(values_only=True), 1):
                if not any(row):  # Skip empty rows
                    continue
                
                if headers is None:
                    # First non-empty row is headers
                    headers = [self._normalize_hebrew_text(str(cell) if cell else "") for cell in row]
                    continue
                
                # Map row to record
                record = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        record[headers[col_idx]] = self._process_cell_value(value)
                
                if record:
                    data_records.append(record)
                    result.records_processed += 1
            
            # Validate data quality
            result.quality_metrics = await self.validate_data_quality(data_records)
            
            # Process in chunks
            for i in range(0, len(data_records), self.chunk_size):
                chunk = data_records[i:i + self.chunk_size]
                chunk_result = await self._process_data_chunk(chunk, i // self.chunk_size + 1)
                
                result.records_imported += chunk_result.records_imported
                result.records_skipped += chunk_result.records_skipped
                result.records_failed += chunk_result.records_failed
                result.errors.extend(chunk_result.errors)
                result.warnings.extend(chunk_result.warnings)
            
            result.status = ProcessingStatus.COMPLETED
            
        except Exception as e:
            logger.error(f"Excel processing failed: {str(e)}")
            result.status = ProcessingStatus.FAILED
            result.errors.append(str(e))
        
        return result
    
    async def _process_csv_file(self, file_path: str, file_hash: str) -> ProcessingResult:
        """Process CSV file with encoding detection and Hebrew support."""
        result = ProcessingResult(
            status=ProcessingStatus.PROCESSING,
            records_processed=0,
            records_imported=0,
            records_skipped=0,
            records_failed=0,
            quality_metrics=DataQualityMetrics(),
            file_hash=file_hash
        )
        
        try:
            # Detect encoding
            encoding = await self._detect_encoding(file_path)
            
            # Read CSV with proper encoding
            data_records = []
            
            async with aiofiles.open(file_path, 'r', encoding=encoding) as f:
                async for row in aiocsv.AsyncDictReader(f):
                    # Process CSV record
                    record = {}
                    for key, value in row.items():
                        if key and value:
                            record[self._normalize_hebrew_text(key)] = self._process_cell_value(value)
                    
                    if record:
                        data_records.append(record)
                        result.records_processed += 1
            
            # Validate data quality
            result.quality_metrics = await self.validate_data_quality(data_records)
            
            # Process in chunks
            for i in range(0, len(data_records), self.chunk_size):
                chunk = data_records[i:i + self.chunk_size]
                chunk_result = await self._process_data_chunk(chunk, i // self.chunk_size + 1)
                
                result.records_imported += chunk_result.records_imported
                result.records_skipped += chunk_result.records_skipped
                result.records_failed += chunk_result.records_failed
                result.errors.extend(chunk_result.errors)
                result.warnings.extend(chunk_result.warnings)
            
            result.status = ProcessingStatus.COMPLETED
            
        except Exception as e:
            logger.error(f"CSV processing failed: {str(e)}")
            result.status = ProcessingStatus.FAILED
            result.errors.append(str(e))
        
        return result
    
    async def _process_json_file(self, file_path: str, file_hash: str) -> ProcessingResult:
        """Process JSON file with Hebrew text support."""
        result = ProcessingResult(
            status=ProcessingStatus.PROCESSING,
            records_processed=0,
            records_imported=0,
            records_skipped=0,
            records_failed=0,
            quality_metrics=DataQualityMetrics(),
            file_hash=file_hash
        )
        
        try:
            # Load JSON data
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                content = await f.read()
                data = json.loads(content)
            
            # Handle different JSON structures
            if isinstance(data, dict):
                if 'data' in data:
                    data_records = data['data']
                elif 'records' in data:
                    data_records = data['records']
                else:
                    data_records = [data]
            elif isinstance(data, list):
                data_records = data
            else:
                raise ValueError("Invalid JSON structure")
            
            result.records_processed = len(data_records)
            
            # Validate data quality
            result.quality_metrics = await self.validate_data_quality(data_records)
            
            # Process in chunks
            for i in range(0, len(data_records), self.chunk_size):
                chunk = data_records[i:i + self.chunk_size]
                chunk_result = await self._process_data_chunk(chunk, i // self.chunk_size + 1)
                
                result.records_imported += chunk_result.records_imported
                result.records_skipped += chunk_result.records_skipped
                result.records_failed += chunk_result.records_failed
                result.errors.extend(chunk_result.errors)
                result.warnings.extend(chunk_result.warnings)
            
            result.status = ProcessingStatus.COMPLETED
            
        except Exception as e:
            logger.error(f"JSON processing failed: {str(e)}")
            result.status = ProcessingStatus.FAILED
            result.errors.append(str(e))
        
        return result
    
    async def _process_data_chunk(self, chunk: List[Dict[str, Any]], chunk_number: int) -> ProcessingResult:
        """Process a chunk of data records."""
        result = ProcessingResult(
            status=ProcessingStatus.PROCESSING,
            records_processed=len(chunk),
            records_imported=0,
            records_skipped=0,
            records_failed=0,
            quality_metrics=DataQualityMetrics()
        )
        
        try:
            for record in chunk:
                try:
                    # Skip if already processed (based on hash)
                    record_hash = self._calculate_record_hash(record)
                    if record_hash in self._processed_hashes:
                        result.records_skipped += 1
                        continue
                    
                    # Map to database model
                    inspection_data = await self._map_to_inspection_model(record)
                    
                    if inspection_data:
                        # Create inspection record
                        inspection = Inspection(**inspection_data)
                        self.session.add(inspection)
                        
                        self._processed_hashes.add(record_hash)
                        result.records_imported += 1
                    else:
                        result.records_skipped += 1
                
                except Exception as e:
                    logger.error(f"Failed to process record: {str(e)}")
                    result.records_failed += 1
                    result.errors.append(f"Record processing failed: {str(e)}")
            
            # Commit the chunk
            await self.session.commit()
            result.status = ProcessingStatus.COMPLETED
            
        except Exception as e:
            await self.session.rollback()
            logger.error(f"Chunk processing failed: {str(e)}")
            result.status = ProcessingStatus.FAILED
            result.errors.append(str(e))
        
        return result
    
    async def _detect_encoding(self, file_path: str) -> str:
        """Detect file encoding."""
        async with aiofiles.open(file_path, 'rb') as f:
            raw_data = await f.read(10000)  # Read first 10KB
        
        detected = chardet.detect(raw_data)
        encoding = detected.get('encoding', 'utf-8')
        
        # Hebrew files are often in cp1255 or utf-8
        if encoding.lower() in ['windows-1255', 'cp1255']:
            return 'cp1255'
        elif encoding.lower() in ['utf-8', 'utf8']:
            return 'utf-8'
        else:
            # Default to utf-8 for safety
            return 'utf-8'
    
    def _process_cell_value(self, value: Any) -> Any:
        """Process cell value with Hebrew text handling."""
        if value is None:
            return None
        
        if isinstance(value, str):
            # Clean and normalize Hebrew text
            cleaned = self._clean_hebrew_text(value)
            normalized = self._normalize_hebrew_text(cleaned)
            return normalized if normalized else None
        
        # Handle dates
        if isinstance(value, (datetime, date)):
            return value
        
        # Handle numbers
        if isinstance(value, (int, float)):
            return value
        
        # Convert other types to string and process
        return self._process_cell_value(str(value))
    
    def _clean_hebrew_text(self, text: str) -> str:
        """Clean Hebrew text from common issues."""
        if not text:
            return text
        
        # Remove BOM and other invisible characters
        text = text.replace('\ufeff', '')  # BOM
        text = text.replace('\u200f', '')  # Right-to-left mark
        text = text.replace('\u200e', '')  # Left-to-right mark
        
        # Normalize whitespace
        text = ' '.join(text.split())
        
        # Remove problematic characters
        text = text.replace('"', '"').replace('"', '"')
        text = text.replace(''', "'").replace(''', "'")
        
        return text.strip()
    
    def _normalize_hebrew_text(self, text: str) -> str:
        """Normalize Hebrew text for consistent processing."""
        if not text:
            return text
        
        # Unicode normalization
        text = unicodedata.normalize('NFD', text)
        
        # Remove Hebrew vowel points (nikkud)
        text = self.HEBREW_POINTS.sub('', text)
        
        # Normalize common Hebrew ligatures
        text = text.replace('\u05f0', '\u05d5\u05d5')  # װ -> וו
        text = text.replace('\u05f1', '\u05d5\u05d9')  # ױ -> וי
        text = text.replace('\u05f2', '\u05d9\u05d9')  # ײ -> יי
        
        return text.strip()
    
    async def _validate_hebrew_fields(self, record: Dict[str, Any]) -> List[str]:
        """Validate Hebrew text fields in record."""
        issues = []
        
        hebrew_fields = [
            'building_manager', 'red_team', 'inspection_type',
            'inspection_leader', 'inspection_notes'
        ]
        
        for field in hebrew_fields:
            value = record.get(field)
            if isinstance(value, str) and value:
                # Check if Hebrew text is present where expected
                if not self.HEBREW_LETTERS.search(value):
                    issues.append(f"Field '{field}' should contain Hebrew text")
                
                # Check for encoding issues
                if '�' in value:
                    issues.append(f"Field '{field}' has encoding issues")
                
                # Check for mixed RTL/LTR issues
                if self._has_mixed_direction_issues(value):
                    issues.append(f"Field '{field}' has mixed text direction issues")
        
        return issues
    
    async def _validate_data_formats(self, record: Dict[str, Any]) -> List[str]:
        """Validate data formats in record."""
        issues = []
        
        # Validate dates
        date_fields = ['execution_schedule', 'target_completion', 'distribution_date']
        for field in date_fields:
            value = record.get(field)
            if value and not isinstance(value, (datetime, date)):
                issues.append(f"Field '{field}' should be a valid date")
        
        # Validate boolean fields
        bool_fields = ['coordinated_with_contractor', 'report_distributed', 'repeat_inspection']
        for field in bool_fields:
            value = record.get(field)
            if value is not None and not isinstance(value, bool):
                if isinstance(value, str):
                    if value.lower() not in ['true', 'false', 'כן', 'לא', 'yes', 'no']:
                        issues.append(f"Field '{field}' should be a valid boolean value")
        
        return issues
    
    def _has_mixed_direction_issues(self, text: str) -> bool:
        """Check if text has mixed RTL/LTR direction issues."""
        hebrew_chars = len(self.HEBREW_LETTERS.findall(text))
        latin_chars = len(re.findall(r'[a-zA-Z]', text))
        
        # If text has both Hebrew and Latin characters, it might have direction issues
        return hebrew_chars > 0 and latin_chars > 0 and len(text) > 10
    
    async def _map_to_inspection_model(self, record: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Map record to inspection model format."""
        try:
            # Basic field mapping
            mapped_data = {
                'building_id': record.get('building_id') or record.get('מבנה'),
                'building_manager': record.get('building_manager') or record.get('מנהל מבנה'),
                'red_team': record.get('red_team') or record.get('צוות אדום'),
                'inspection_type': record.get('inspection_type') or record.get('סוג הבדיקה'),
                'inspection_leader': record.get('inspection_leader') or record.get('מוביל הבדיקה'),
                'inspection_round': self._safe_int(record.get('inspection_round') or record.get('סבב בדיקות')),
                'regulator_1': record.get('regulator_1') or record.get('רגולטור 1'),
                'regulator_2': record.get('regulator_2') or record.get('רגולטור 2'),
                'regulator_3': record.get('regulator_3') or record.get('רגולטור 3'),
                'regulator_4': record.get('regulator_4') or record.get('רגולטור 4'),
                'execution_schedule': self._safe_date(record.get('execution_schedule') or record.get('לוז ביצוע')),
                'target_completion': self._safe_date(record.get('target_completion') or record.get('יעד לסיום')),
                'coordinated_with_contractor': self._safe_bool(record.get('coordinated_with_contractor') or record.get('מתואם מול זכיין')),
                'defect_report_path': record.get('defect_report_path') or record.get('נתיב דוח ליקויים'),
                'report_distributed': self._safe_bool(record.get('report_distributed') or record.get('דוח הופץ')),
                'distribution_date': self._safe_date(record.get('distribution_date') or record.get('תאריך הפצה')),
                'repeat_inspection': self._safe_bool(record.get('repeat_inspection') or record.get('בדיקה חוזרת')),
                'inspection_notes': record.get('inspection_notes') or record.get('התרשמות'),
            }
            
            # Remove None values
            mapped_data = {k: v for k, v in mapped_data.items() if v is not None}
            
            # Validate required fields
            if not mapped_data.get('building_id'):
                return None
            
            return mapped_data
            
        except Exception as e:
            logger.error(f"Failed to map record: {str(e)}")
            return None
    
    def _safe_int(self, value: Any) -> Optional[int]:
        """Safely convert value to integer."""
        if value is None:
            return None
        
        try:
            if isinstance(value, int):
                return value
            elif isinstance(value, float):
                return int(value)
            elif isinstance(value, str):
                cleaned = value.strip()
                if not cleaned:
                    return None
                return int(float(cleaned))
        except (ValueError, TypeError):
            pass
        
        return None
    
    def _safe_date(self, value: Any) -> Optional[date]:
        """Safely convert value to date."""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, date):
            return value
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try common date formats
            formats = ['%Y-%m-%d', '%d/%m/%Y', '%d.%m.%Y', '%Y/%m/%d']
            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        return None
    
    def _safe_bool(self, value: Any) -> bool:
        """Safely convert value to boolean."""
        if value is None:
            return False
        
        if isinstance(value, bool):
            return value
        
        if isinstance(value, str):
            value = value.strip().lower()
            return value in ['true', '1', 'yes', 'כן', 'v', '✓']
        
        return bool(value)
    
    async def get_processing_statistics(self) -> Dict[str, Any]:
        """Get processing statistics and metrics."""
        return {
            'processed_files': len(self._processed_hashes),
            'unique_records': len(self._processed_hashes),
            'last_processing_time': datetime.now().isoformat(),
            'supported_formats': [format.value for format in DataFormat if format != DataFormat.UNKNOWN],
            'chunk_size': self.chunk_size,
        }